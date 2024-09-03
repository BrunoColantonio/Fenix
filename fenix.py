import flet as ft
import sqlite3
import math
import pandas
import sys
import os
import json

import openpyxl
from openpyxl.styles import Font, Alignment

from fpdf import FPDF, XPos, YPos

from datetime import datetime

#Auxiliary pyinstaller function 
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#GLOBALs
CURRENT_USER = ""
PENDING_FILENAME = ""
DATABASE_PATH = resource_path("data\\fenix.db")
FONT = "ModeRustic"
#MACROS
IVA = 1.21
A_PERCENT = 0.3
N_PERCENT = 0.7
LIMIT_OF_ELEMENTS = 500

#IDXs
CANT_A_IDX = 0
CANT_N_IDX = 1
CANT_TOTAL_IDX = 2
CODE_IDX = 3
PRODUCT_IDX = 4
UNITARY_PRICE_IDX = 5
A_PRICE_IDX = 6
N_PRICE_IDX = 7
CLIENT_IDX = 0
ZONE_IDX = 1
TIME_IDX = 2

#COLORS
RED = "#b03a2e"
GREY = "#aeb6bf"
DISABLED_GREY = "#404447"

BG_COLOR = "#222222"
BUTTONS_BORDER_COLOR = "#ecf0f1"
LIST_PRESSED_FILL_COLOR = "#b03a2e"
PRICE_BUTTONS_COLOR = "#b03a2e"
CLEAR_CLIENT_BUTTON_COLOR = "#b03a2e"
SWITCH_COLOR = "#b03a2e"
CURSOR_COLOR = "#b03a2e"
HINT_TEXT_COLOR = "#b03a2e"
ADD_BUTTON_BGCOLOR = GREY
ADD_BUTTON_COLOR = RED
UPDATE_BUTTON_BGCOLOR = GREY
UPDATE_BUTTON_COLOR = RED

# DROPDOWN_BG_COLOR = "#17202a"
DROPDOWN_BG_COLOR = "#2d2f2e"

ERROR_MSG_COLOR = "#b03a2e"
INVALID_CODE_MSG_COLOR = "#b03a2e"
DELETED_MSG_COLOR = "#b03a2e"

UPDATED_MSG_COLOR = "#aeb6bf"
UPDATING_LIST_MSG_COLOR = "#aeb6bf"
ADDED_CLIENT_MSG_COLOR = "green"
UPDATED_CLIENT_MG_COLOR = GREY

CREATED_ORDER_MSG_COLOR = "green"
HOME_ICON_COLOR = "#82e0aa"

#LISTS
users_list = []
modes = ["F1","F2","F3"]
table_headers = ["A","N","Cant. Total","Código","Descripción","Precio Unitario","Total A","Total N"]
product_list = []
client_list = []

#FUNCTIONS

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def set_clients(user, page):
    global client_list
    client_list.clear()
    conn = sqlite3.connect(DATABASE_PATH)
    query = f"SELECT Cliente FROM Cliente INNER JOIN Usuario ON Cliente.Usuario = Usuario.ID WHERE Usuario.Usuario = '{user}'"

    try:
        cursor = conn.execute(query)
        clients = cursor.fetchall()

        if(clients == None):
                show_error_message(page)
        else:
            for client in clients:
                client_list.append(client)
    except:
        show_error_message(page)
    finally:
        conn.close()

def set_users_list(page):
        #If Fenix:
        conn = sqlite3.connect(DATABASE_PATH)
        query = f"SELECT Usuario FROM Usuario"
        try:
            cursor = conn.execute(query)
            users = cursor.fetchall()

            for user in users:
                users_list.append(user[0])
        except:
            show_error_message(page)
        finally:
            conn.close()

        # If Martin:
        # users_list.append("Martin")

        # If Emma:
        # users_list.append("Emmanuel")

        # # If Nadia:
        # users_list.append("Nadia")

def set_products(page):
        global product_list
        product_list.clear()
        conn = sqlite3.connect(DATABASE_PATH)
        query = "SELECT Descripcion FROM Producto"
        try:
            cursor = conn.execute(query)
            products = cursor.fetchall()

            if(products == None):
                    show_error_message(page)
            else:
                for product in products:
                    product_list.append(product)
        except:
            show_error_message(page)
        finally:
            conn.close()

def set_dropdown_options(list):
    options = []
    for item in list:
        options.append(ft.dropdown.Option(item))
    return options

def show_error_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("HA OCURRIDO UN ERROR.", size = 20, color=GREY),
        bgcolor = ERROR_MSG_COLOR,
        duration=600)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior

    page.update()

def show_invalid_code_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("NO EXISTE UN PRODUCTO CON EL CÓDIGO INGRESADO.", size = 20, color=GREY),
        bgcolor = INVALID_CODE_MSG_COLOR,
        duration=850)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior

    page.update()

def show_updated_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("SE HA ACTUALIZADO EL DATO CORRECTAMENTE.", size = 20,color=RED),
        bgcolor = UPDATED_MSG_COLOR,
        duration=600)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_deleted_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("SE HA ELIMINADO EL DATO CORRECTAMENTE.", size = 20, color=GREY),
        bgcolor = DELETED_MSG_COLOR,
        duration=600)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_created_order_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Row(controls=[ft.Text("¡Pedido generado correctamente!", size = 20,font_family=FONT,color="black"),ft.Icon(ft.icons.THUMB_UP,color="black")]),
        bgcolor = CREATED_ORDER_MSG_COLOR,
        duration=1300)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_created_budget_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Row(controls=[ft.Text("¡Presupuesto generado correctamente!", size = 20,font_family=FONT,color="black"),ft.Icon(ft.icons.THUMB_UP,color="black")]),
        bgcolor = CREATED_ORDER_MSG_COLOR,
        duration=1300)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_generating_pdf_message(page):
    snack_bar = ft.SnackBar(content=ft.Column(controls=[
                ft.Text("Generando presupuesto", size = 20, font_family=FONT),
                ft.ProgressBar(width=1500,color=RED, bgcolor="#eeeeee")]),
                bgcolor = UPDATING_LIST_MSG_COLOR,
                duration=10000)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    page.update()

def show_updating_list_message(page):
    snack_bar = ft.SnackBar(content=ft.Column(controls=[
                ft.Text("Actualizando lista de precios", size = 20, font_family=FONT),
                ft.ProgressBar(width=1500,color=RED, bgcolor="#eeeeee")]),
                bgcolor = UPDATING_LIST_MSG_COLOR,
                duration=17000)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    page.update()

def show_successfully_updated_list_message(page):
    snack_bar = ft.SnackBar(ft.Text("Lista actualizada correctamente!!", size = 20, font_family=FONT),
                                 bgcolor = UPDATED_MSG_COLOR,
                                 duration=1300)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    page.update()

def show_empty_fields_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("DEBE COMPLETAR TODOS LOS DATOS", size = 20,color=GREY,font_family=FONT),
        bgcolor = DELETED_MSG_COLOR,
        duration=850)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_added_client_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("CLIENTE AÑADIDO CORRECTAMENTE", size = 20,color="black",font_family=FONT),
        bgcolor = ADDED_CLIENT_MSG_COLOR,
        duration=600)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_updated_client_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("CLIENTE ACTUALIZADO CORRECTAMENTE", size = 20,color=RED,font_family=FONT),
        bgcolor = UPDATED_CLIENT_MG_COLOR,
        duration=600)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_deleted_client_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("CLIENTE ELIMINADO CORRECTAMENTE", size = 20,color=GREY,font_family=FONT),
        bgcolor = RED,
        duration=600)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

def show_duplicated_product(page,product):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text(f"YA SE HA INTRODUCIDO EL PRODUCTO '{product}'", size = 20, color=GREY),
        bgcolor = INVALID_CODE_MSG_COLOR,
        duration=850)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior

    page.update()

def show_added_pending_message(page):
    #MENSAJE con Barra inferior
    snack_bar = ft.SnackBar(
        ft.Text("PENDIENTES AÑADIDOS CORRECTAMENTE", size = 20,color="black",font_family=FONT),
        bgcolor = ADDED_CLIENT_MSG_COLOR,
        duration=600)
    page.overlay.append(snack_bar)
    snack_bar.open = True
    #FIN DEL MENSAJE inferior
    page.update()

#SCREENS

class Form(ft.Container):

    def __init__(self, page: ft.Page):
        super().__init__(expand = True)

        set_products(page)
        self.order = []
        self.current_checkbox_selection = ""

        self.page = page
        self.search_mode = "Product"

        self.create_widgets()

        #Price variables
        self.N_SUBTOTAL = 0
        self.A_SUBTOTAL = 0
        self.TOTAL_PRICE = 0
        
     
        #FORM
        self.form = ft.Container(bgcolor = BG_COLOR,
                                 border_radius = 10,
                                 padding = 10,
                                 col = 3,
                                content = ft.Column(
                                    controls = [
                                        self.title,
                                        ft.ResponsiveRow(controls=[self.client,self.clear_client_button],vertical_alignment=ft.CrossAxisAlignment.CENTER),
                                        ft.Stack([self.mode, self.client_search_list]),
                                        self.zone,
                                        self.order_title,
                                        self.search_switch,
                                        # ft.Row(controls=[self.sin_cargo_chkbox,self.descontar_chkbox,self.facturar_chkbox],
                                        #        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                                        #        alignment=ft.MainAxisAlignment.CENTER),
                                        ft.ResponsiveRow(controls=[
                                            ft.Column(controls=[self.sin_cargo_chkbox], col={"xs": 12, "sm": 6, "md": 4, "lg": 4}),
                                            ft.Column(controls=[self.descontar_chkbox], col={"xs": 12, "sm": 6, "md": 4, "lg": 4}),
                                            ft.Column(controls=[self.facturar_chkbox], col={"xs": 12, "sm": 6, "md": 4, "lg": 4}),
                                        ]),
                                        # self.product,
                                        # self.code,
                                        # self.search_list,
                                        # self.quantity,
                                        self.code,
                                        self.product,
                                        ft.Stack([self.quantity, self.search_list]),
                                        #self.add_button
                                        ft.Row(
                                            controls = [
                                                self.add_button,
                                                self.update_button
                                                ],
                                                alignment=ft.MainAxisAlignment.CENTER
                                            )
                                        
                                    ])
                                )

        #TABLE
        self.table = ft.Container(bgcolor = BG_COLOR,
                                  border_radius = 10,
                                  col = 9,
                                  content = 
                                            ft.Column(
                                                expand = True,
                                                scroll = "auto",
                                                controls = [
                                                    ft.ResponsiveRow([
                                                        self.data_table
                                                    ]
                                                    ),
                                                    ft.Row(controls = [
                                                        self.N_SUBTOTAL_container,
                                                        self.A_SUBTOTAL_container,
                                                        self.TOTAL_PRICE_container
                                                    ],
                                                    alignment=ft.MainAxisAlignment.CENTER),

                                                    ft.Container(content=ft.Row(
                                                        controls=[
                                                            self.first_order_chkbox],
                                                            alignment=ft.MainAxisAlignment.CENTER)),
                                                    ft.Container(content=ft.Row(
                                                        controls=[
                                                            self.generate_order_button,
                                                            self.generate_budget_button],
                                                            alignment=ft.MainAxisAlignment.CENTER)),
                                                    ft.Container(content=ft.Row(
                                                        controls=[
                                                            self.pending_file,
                                                            self.delete_pending_file],
                                                            alignment=ft.MainAxisAlignment.CENTER))

                                                ]
                                            )
                                        
                                    )         

        #Screen view
        self.screen = ft.Container(bgcolor = BG_COLOR,
                                   expand=1,
                                    content=ft.Column(
                                        controls = [
                                            ft.Container(content = ft.ResponsiveRow(controls=[self.form,self.table]),expand=True)
                                        ]))

        self.content = ft.View(
                "/home_screen",
                [
                    self.appbar,
                    self.screen,
                ],
                bgcolor=BG_COLOR
            )
        
        self.content.controls.append(self.pick_files_dialog)
        self.content.controls.append(self.pick_pending_file_dialog)
        self.content.controls.append(self.pick_excel_path_dialog)
        self.content.controls.append(self.pick_excel_filename_dialog)
        self.content.controls.append(self.pick_pdf_path_dialog)

    def build(self):
        return self.content
  
    def change_bg_color(self, e):
        if(e.control.bgcolor == BG_COLOR):
            e.control.bgcolor = DISABLED_GREY
        else:
            e.control.bgcolor = BG_COLOR
        self.page.update()

    def create_widgets(self):
        global users_list

        #DIALOGS
        self.pick_files_dialog = ft.FilePicker(on_result=self.update_price_list)
        self.page.add(self.pick_files_dialog)

        self.pick_pending_file_dialog = ft.FilePicker(on_result=self.add_pending)
        self.page.add(self.pick_pending_file_dialog)

        self.pick_excel_path_dialog = ft.FilePicker(on_result=self.generate_excel)
        self.page.add(self.pick_excel_path_dialog)

        self.pick_excel_filename_dialog = ft.FilePicker(on_result=self.generate_excel)
        self.page.add(self.pick_excel_filename_dialog)
        
        self.pick_pdf_path_dialog = ft.FilePicker(on_result=self.generate_pdf)
        self.page.add(self.pick_pdf_path_dialog)


        # HEADER WIDGETS
        self.user = ft.Dropdown(label = "Usuario",
                                bgcolor=DROPDOWN_BG_COLOR,
                                fill_color=BG_COLOR,
                                text_style=ft.TextStyle(font_family=FONT,size=20),
                                label_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                                suffix_icon=ft.icons.PERSON,
                                options = set_dropdown_options(users_list),
                                border_color = BUTTONS_BORDER_COLOR,
                                on_change = self.set_user)
        
        self.welcome_title = ft.Text("SELECCIONE USUARIO", font_family=FONT,size=30,color="white")

        self.fenix_logo = ft.Image(src=resource_path("assets\\fenix_logo.png"),height=60,width=60,color=RED)

        self.appbar = ft.AppBar(
        leading=ft.Icon(ft.icons.HOME, color=HOME_ICON_COLOR),
        leading_width=40,
        center_title=True,
        title=ft.Container(ft.Row(controls=[self.fenix_logo,self.welcome_title],
                                  alignment=ft.MainAxisAlignment.CENTER)),
                                  bgcolor=BG_COLOR,
                                  actions=[
                                      self.user,
                                        ft.IconButton(ft.icons.UPLOAD_FILE_ROUNDED,tooltip="Actualizar lista de precios",
                                                    on_click=lambda _: self.pick_files_dialog.pick_files(file_type=ft.FilePickerFileType.CUSTOM,
                                                                                                         allowed_extensions=["xlsx", "xls"]),
                                                    icon_color="green"),
                                        ft.IconButton(ft.icons.PENDING_ACTIONS,tooltip="Añadir pendientes al pedido",
                                                    on_click=lambda _: self.pick_pending_file_dialog.pick_files(file_type=ft.FilePickerFileType.CUSTOM,
                                                                                                         allowed_extensions=["xlsx", "xls"]),
                                                    icon_color="yellow,90"),
                                        ft.IconButton(ft.icons.PEOPLE,tooltip="Clientes",
                                                    on_click=self.go_to_clients_screen,
                                                    icon_color=RED)
                                            ]
                                        )

        #FORM WIDGETS
        self.title = ft.Container(ft.Text("Cargue pedido",text_align=ft.TextAlign.CENTER,
                                          size = 30, font_family=FONT, color="white"),
                                  alignment=ft.alignment.center)

        self.client = ft.TextField(label = "Cliente",
                                   label_style=ft.TextStyle(color="white",font_family=FONT),
                                   text_style=ft.TextStyle(color="white",font_family=FONT,size=15),
                                   col=10,
                                   border_color = BUTTONS_BORDER_COLOR,
                                   cursor_color= CURSOR_COLOR,
                                   suffix_icon = ft.icons.SEARCH,
                                   read_only=True,
                                   on_change = self.find_client,
                                   on_focus=self.check_user)
        
        self.clear_client_button = ft.IconButton(icon=ft.icons.HIGHLIGHT_REMOVE_SHARP,
                                                 icon_color=CLEAR_CLIENT_BUTTON_COLOR,
                                                 on_click=self.reset_client,
                                                 col=2)

        self.client_search_list = ft.ListView(visible = False,
                                       divider_thickness = 1,
                                        padding = 10,
                                        spacing = 5,
                                        controls = [],
                                        auto_scroll=False,
                                        expand=True,
                                        height=100)

        self.mode = ft.Dropdown(label = "Modo",
                                text_style=ft.TextStyle(color="white",font_family=FONT),
                                label_style=ft.TextStyle(color="white",font_family=FONT),
                                options = set_dropdown_options(modes),
                                value="F1",
                                border_color = BUTTONS_BORDER_COLOR,
                                bgcolor=DROPDOWN_BG_COLOR,
                                fill_color=BG_COLOR,
                                on_change = self.update_table)
        
        self.zone = ft.TextField(label="Zona",
                                 text_style=ft.TextStyle(color="white",font_family=FONT),
                                 label_style=ft.TextStyle(color="white",font_family=FONT),
                                 border_color=BUTTONS_BORDER_COLOR,
                                 bgcolor=BG_COLOR,
                                 read_only=True)
        
        self.order_title = ft.Container(ft.Text("Añada productos", text_align = "center",
                                                size = 30, font_family=FONT, color="white"),
                                                alignment=ft.alignment.center)

        self.search_switch = ft.Container(content=ft.Switch(label = "Buscar por código",
                                                            label_style=ft.TextStyle(color="white",font_family=FONT),
                                                            label_position=ft.LabelPosition.LEFT,
                                                            value = False,
                                                            scale=0.97,
                                                            inactive_track_color=GREY,
                                                            active_color=SWITCH_COLOR,
                                                            on_change = self.switch_browse),
                                                        alignment=ft.alignment.center)

        #Checkboxs
        self.sin_cargo_chkbox = ft.Checkbox(label="Sin cargo", value=False,fill_color=GREY, check_color=RED, on_change=self.check_checkbox,label_style=ft.TextStyle(color="white",font_family=FONT))
        self.descontar_chkbox = ft.Checkbox(label="Descontar", value=False,fill_color=GREY, check_color=RED, on_change=self.check_checkbox,label_style=ft.TextStyle(color="white",font_family=FONT))
        self.facturar_chkbox = ft.Checkbox(label="Facturar", value=False,fill_color=GREY, check_color=RED, on_change=self.check_checkbox,label_style=ft.TextStyle(color="white",font_family=FONT))

        self.product = ft.TextField(label = "Producto",
                                    label_style=ft.TextStyle(color="white",font_family=FONT),
                                    text_style=ft.TextStyle(color="white",font_family=FONT),
                                    cursor_color=CURSOR_COLOR,
                                    border_color = BUTTONS_BORDER_COLOR,
                                    bgcolor=BG_COLOR,
                                    suffix_icon = ft.icons.SEARCH,
                                    on_change = self.find_product)
        
        self.code = ft.TextField(label = "Código",
                                 label_style=ft.TextStyle(color="white",font_family=FONT),
                                text_style=ft.TextStyle(color="white",font_family=FONT),
                                border_color = BUTTONS_BORDER_COLOR,
                                cursor_color=CURSOR_COLOR,
                                bgcolor=BG_COLOR,
                                suffix_icon = ft.icons.SEARCH,
                                visible = False,
                                on_change=self.check_code_input,
                                on_submit=lambda _:self.quantity.focus())
        
        self.code.on_key_down = self.handle_code_key_event
        
        self.search_list = ft.ListView(visible = False,
                                       divider_thickness = 1,
                                        padding = 10,
                                        spacing = 5,
                                        controls = [],
                                        auto_scroll=False,
                                        expand=True,
                                        height=400)

        self.quantity = ft.TextField(label = "Cantidad",
                                     label_style=ft.TextStyle(color="white",font_family=FONT),
                                     text_style=ft.TextStyle(color="white",font_family=FONT),
                                     cursor_color=CURSOR_COLOR,
                                     bgcolor=BG_COLOR,
                                     border_color = BUTTONS_BORDER_COLOR,
                                     read_only=True,
                                     input_filter = ft.NumbersOnlyInputFilter(),
                                     on_change=self.check_input,
                                     on_submit=self.check_submit,
                                     on_focus=self.check_quantity_input)

        self.add_button = ft.ElevatedButton(text = "Añadir",
                                            on_click = self.add_product,
                                            icon=ft.icons.ADD,
                                            disabled = True,
                                            style=ft.ButtonStyle(color={"":ADD_BUTTON_COLOR,"disabled":BG_COLOR},
                                                                 bgcolor={"":ADD_BUTTON_BGCOLOR,"disabled":DISABLED_GREY}))

        self.update_button = ft.ElevatedButton(text = "Actualizar", on_click = self.update_product,
                                               icon=ft.icons.REFRESH,
                                               disabled=True,
                                               style=ft.ButtonStyle(color={"":ADD_BUTTON_COLOR,"disabled":BG_COLOR},
                                                                 bgcolor={"":ADD_BUTTON_BGCOLOR,"disabled":DISABLED_GREY}))

        #TABLE WIDGETS
        self.data_table = ft.DataTable(expand = True,
                                       border = ft.border.all(2,BUTTONS_BORDER_COLOR),
                                        column_spacing=10,
                                        border_radius = 10,
                                        columns = [
                                             ft.DataColumn(label=ft.Container(ft.Text("Cant. A",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=55)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Cant. N",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=60)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Cant. Total",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=50)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Código",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=50)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Descripción",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=100)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Precio unitario",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=80)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Total A",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=45)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Total N",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=45)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Acciones",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT),
                                                                              width = 70, expand=True, alignment=ft.alignment.center))

                                        ],
                                        rows = [])
        
        self.N_SUBTOTAL_text = ft.Text(f"Precio en negro: $0",font_family=FONT,color="white")
        self.A_SUBTOTAL_text = ft.Text(f"Precio en blanco: $0",font_family=FONT,color="white")
        self.TOTAL_PRICE_text = ft.Text(f"Precio total: $0",font_family=FONT,color="white")

        self.N_SUBTOTAL_container = ft.Container(content=self.N_SUBTOTAL_text,border_radius=5,bgcolor=PRICE_BUTTONS_COLOR,padding=5)
        self.A_SUBTOTAL_container = ft.Container(content=self.A_SUBTOTAL_text,border_radius=5,bgcolor=PRICE_BUTTONS_COLOR,padding=5)
        self.TOTAL_PRICE_container = ft.Container(content=self.TOTAL_PRICE_text,border_radius=5,bgcolor=PRICE_BUTTONS_COLOR,padding=5)

        # desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.first_order_chkbox = ft.Checkbox(label=f"Primer pedido del día para el cliente ''",
                                              value=True,fill_color=GREY, check_color=RED,
                                              label_style=ft.TextStyle(color="white",font_family=FONT))

        self.generate_order_button = ft.ElevatedButton("Generar pedido",bgcolor="green",
                                                       color="white",
                                                       disabled=True,
                                                       icon=ft.icons.DOWNLOADING_ROUNDED,
                                                       on_click=lambda _:self.pick_excel_path_dialog.get_directory_path() if self.first_order_chkbox.value else self.pick_excel_filename_dialog.save_file())
        
        self.generate_budget_button = ft.ElevatedButton("Generar presupuesto",bgcolor="blue",
                                                       color="white",
                                                       disabled=True,
                                                       icon=ft.icons.PICTURE_AS_PDF,
                                                       on_click=lambda _:self.pick_pdf_path_dialog.get_directory_path())
        
        self.pending_file = ft.Text("",visible=False,style=ft.TextStyle(color="white",font_family=FONT))
        self.delete_pending_file = ft.IconButton(icon=ft.icons.REMOVE_CIRCLE, tooltip="Eliminar pendiente", visible=False,
                                                 on_click=self.delete_pending, icon_color=RED, icon_size=20)

    def handle_code_key_event(self, e: ft.KeyboardEvent):
        if (e.key == "Tab"):
            self.quantity.focus()
        self.page.update()

    def check_user(self, e):
        if(CURRENT_USER == ""):
            self.client.error_text = "Primero debe seleccionar usuario" 

        self.client.update()

    def reset_checkboxs(self):
        self.facturar_chkbox.disabled = False
        self.descontar_chkbox.disabled = False
        self.sin_cargo_chkbox.disabled = False

        self.facturar_chkbox.value = ""
        self.descontar_chkbox.value = ""
        self.sin_cargo_chkbox.value = ""

        self.current_checkbox_selection = ""

        self.page.update()

    def check_checkbox(self, e):

        #If it was deselected
        if(e.control.value == False):
            self.reset_checkboxs()
        else:
            self.current_checkbox_selection = e.control.label

            match (self.current_checkbox_selection):
                case "Sin cargo":
                    self.facturar_chkbox.disabled = True
                    self.descontar_chkbox.disabled = True
                case "Facturar":
                    self.descontar_chkbox.disabled = True
                    self.sin_cargo_chkbox.disabled = True
                case "Descontar":
                    self.sin_cargo_chkbox.disabled = True
                    self.facturar_chkbox.disabled = True
            
        self.page.update()

    def check_quantity_input(self, e):
        if(self.product.value == "" and self.search_mode == "Product"):
            self.quantity.read_only = True
            self.quantity.error_text = "Primero debe seleccionar un producto"
        elif(self.code.value == "" and self.search_mode == "Code"):
            self.quantity.read_only = True
            self.quantity.error_text = "Primero debe ingresar un código"
        
        self.quantity.update()

    def switch_browse(self, e):

        self.search_list.visible = False
        self.quantity.visible = True
        self.page.update()

        if(self.search_mode == "Product"):
            self.product.visible = not self.product.visible
            self.code.visible = not self.code.visible
            self.product.value = ""
            self.search_mode = "Code"


        else:
            self.code.visible = False
            self.product.visible = True
            self.code.value = ""
            self.search_mode = "Product"           
        
        self.quantity.error_text = ""
        self.quantity.value = ""
        self.page.update()

    def switch_mode(self, e):
        self.mode.visible = not self.mode.visible
        self.max_mode.visible = not self.max_mode.visible
        self.page.update()

    def check_client_input(self, e):
        if e.control.value == "":
            #Disable product and quantity inputs
            self.product.disabled = True
            self.code.disabled = True
            self.quantity.disabled = True
        else:
            self.product.disabled = True
            self.code.disabled = True
            self.quantity.disabled = True

    def check_input(self,e):
        if(e.control.value == "" or self.update_button.disabled == False):
            self.add_button.disabled = True
        else:
            self.add_button.disabled = False

        self.page.update()

    def check_submit(self, e):
        if(self.add_button.disabled == True):
            self.update_product(e)
        elif(self.update_button.disabled == True):
            self.add_product(e)

    def check_code_input(self, e):
        if e.control.value !=  "":
            self.quantity.error_text = ""
            self.quantity.read_only = False

        self.page.update()
            
    def find_product(self, e):
            search_text = e.control.value.lower()
            self.search_list.clean()
            found = False

            #Auxiliary variable
            number_of_elements = 0
            for product in product_list:
                #if str(product[0].lower()).startswith(search_text) and show_limit > 0:
                if search_text in product[0].lower() and number_of_elements < LIMIT_OF_ELEMENTS:
                #if search_text in product[0].lower():

                    self.search_list.controls.append(ft.Container(on_click = lambda e: self.select_product(e),
                                                                    content = ft.Text(product[0],font_family=FONT,
                                                                                      color="white"),
                                                                                      ink=True,ink_color=LIST_PRESSED_FILL_COLOR,
                                                                                      bgcolor=BG_COLOR,
                                                                                      on_hover=self.change_bg_color))
                    found = True
                    number_of_elements += 1

                    if(number_of_elements == LIMIT_OF_ELEMENTS):
                        break
                    
                #If input is empty, do not display list
                if e.control.value ==  "":
                    self.search_list.visible = False
                    self.quantity.visible = True
                else:
                    self.search_list.visible = True
                    self.quantity.visible = False
                    self.quantity.error_text = ""
                    self.quantity.read_only = False

            if not found:
                self.search_list.controls.append(
                    ft.Container(content = ft.Text("No hay coincidencias",font_family=FONT,color="white",bgcolor=BG_COLOR)))
            #Update screen
            self.page.update()
    
    def find_client(self, e):
            search_text = e.control.value.lower()
            self.client_search_list.clean()
            found = False

            #Auxiliary variable
            for client in client_list:
                if search_text in client[0].lower():

                    self.client_search_list.controls.append(ft.Container(on_click = lambda e: self.select_client(e),
                                                                    content = ft.Text(client[0],font_family=FONT,
                                                                                      color="white"),
                                                                                      ink=True,ink_color=LIST_PRESSED_FILL_COLOR,
                                                                                      bgcolor=BG_COLOR,
                                                                                      on_hover=self.change_bg_color))
                    found = True

                #If input is empty, do not display list
                if e.control.value ==  "":
                    self.client_search_list.visible = False
                    self.mode.visible = True
                else:
                    self.client_search_list.visible = True
                    self.mode.visible = False

            if not found:
                self.client_search_list.controls.append(
                    ft.Container(content = ft.Text("No hay coincidencias",font_family=FONT,color="white",bgcolor=BG_COLOR)))
            #Update screen
            self.page.update()

    def select_product(self,e):
        
        self.selected_product = e.control.content.value
        self.product.value = e.control.content.value
        self.focused_textField = False
        self.search_list.visible = False
        self.quantity.visible = True
        self.quantity.focus()
        self.product.update()
        self.search_list.update()
        self.page.update()

    def select_client(self,e):
        
        selected_client = e.control.content.value
        self.client.value = selected_client
        self.first_order_chkbox.label = f"Primer pedido del día para el cliente '{self.client.value}'"
        self.set_zone(selected_client)
        self.focused_textField = False
        self.client_search_list.visible = False
        self.mode.visible = True

        if(self.generate_order_button.disabled == True and len(self.data_table.rows) != 0):
            self.generate_order_button.disabled = False

        self.mode.focus()
        # self.product.update()
        # self.search_list.update()
        # self.page.update()
        self.page.update()

    def set_zone(self,client):
        conn = sqlite3.connect(DATABASE_PATH)
        query = f"SELECT Zona FROM Cliente WHERE Cliente = '{client}'"
        try:
            cursor = conn.execute(query)
            zone = cursor.fetchall()

            zone = zone[0]
            #self.zone.disabled = False
            self.zone.value = zone[0]
            #self.zone.disabled = True
            self.page.update()
        except:
            show_error_message(self.page)
        finally:
            conn.close()

    def reset_client(self, e):
        self.client.value = ""
        self.first_order_chkbox.label = f"Primer pedido del día para el cliente"
        self.first_order_chkbox.value = True
        self.zone.value = ""
        self.client_search_list.visible = False
        self.mode.visible = True
        self.generate_order_button.disabled = True
        self.generate_budget_button.disabled = True
        
        self.page.update()

    def select_code(self,e):
        
        self.selected_code = e.control.content.value
        self.code.value = e.control.content.value

        self.search_list.visible = False
        self.quantity.visible = True
        self.search_list.update()
        self.page.update()

    def add_product(self,e):
        match(self.mode.value):
            case "F1":
                cant_A = self.quantity.value
                cant_N = 0
                cant_Total = cant_A
            case "F2":
                cant_A = math.floor(int(self.quantity.value) / 2)
                cant_N = math.ceil(int(self.quantity.value) / 2)
                cant_Total = self.quantity.value
            case "F3":
                cant_A = 0
                cant_N = self.quantity.value
                cant_Total = cant_N
        
        #Get product from DB
        if(self.search_mode == "Product"):
            code,product,price = self.get_product("Product",self.product.value)
        else:
            code,product,price = self.get_product("Code",self.code.value)

        #Check checkboxs
        if(self.current_checkbox_selection != ""):
            #If any is checked, set prices to 0 and add it to product description
            product = f"{product} [{self.current_checkbox_selection}]"
            price = 0
            A_price = 0
            N_price = 0

            #Insert into table
            self.insert_table_row(cant_A, cant_N, cant_Total, code, product, price, A_price, N_price)
                    
            self.clear_fields()
            self.add_button.disabled = True

            self.update_price_widgets()
                    
            if(self.generate_order_button.disabled == True):
                self.generate_order_button.disabled = False
                    
            if(self.generate_budget_button.disabled == True and self.mode.value != "F2"):
                self.generate_budget_button.disabled = False

        elif(product in self.order):
            show_duplicated_product(self.page, product)
            self.product.value = ""
            self.code.value = ""
            self.quantity.value = ""
            self.page.update()

        else:
            #If code exists:
            if(code != "not exists"):
                price = round(price,2)
                A_price = float(cant_A) * price * IVA
                N_price = float(cant_N) * price

                A_price = round(A_price,2)
                N_price = round(N_price,2)

                #Update total price variables
                self.A_SUBTOTAL += A_price
                self.N_SUBTOTAL += N_price
                self.TOTAL_PRICE += A_price + N_price

                #Format prices
                price = '{:,.2f}'.format(price)
                A_price = '{:,.2f}'.format(A_price)
                N_price = '{:,.2f}'.format(N_price)

                #Insert into table
                self.insert_table_row(cant_A, cant_N, cant_Total, code, product, price, A_price, N_price)
                    
                self.clear_fields()
                self.add_button.disabled = True

                self.update_price_widgets()
                    
                if(self.generate_order_button.disabled == True):
                    self.generate_order_button.disabled = False
                    
                if(self.generate_budget_button.disabled == True and self.mode.value != "F2"):
                    self.generate_budget_button.disabled = False

        self.page.update()
    
    def update_product(self, e):
        self.add_product(e)
        self.add_button.disabled = False
        self.update_button.disabled = True

        show_updated_message(self.page)

        self.page.update()

    def update_table(self,e):
        # GET ALL DATA TABLE ROWS
        all_rows = []
        special_products = []
        for row in self.data_table.rows:
            row_data = []
            for cell in row.cells:                
                try:
                    row_data.append(cell.content.value)
                except Exception as e:
                    pass

            #If not a special product, append to all rows
            if row_data[UNITARY_PRICE_IDX] != 0:
                all_rows.append(row_data)
            else:
                special_products.append(row_data)
                
        
        #DELETE ALL ROWS
        self.clean_data_table()        
        #UPDATE TABLE WITH NEW MODE
        for row in all_rows:
            self.update_mode_add_product(row[CANT_TOTAL_IDX],row[PRODUCT_IDX])

        if len(special_products) > 0:
            for row in special_products:
                self.current_checkbox_selection = "aux"
                self.insert_table_row(row[CANT_A_IDX],row[CANT_N_IDX],row[CANT_TOTAL_IDX],row[CODE_IDX],row[PRODUCT_IDX],
                                      row[UNITARY_PRICE_IDX],row[A_PRICE_IDX],row[N_PRICE_IDX])
                self.current_checkbox_selection = ""
        
        #CHECK BUDGET BUTTON
        if(self.mode.value == "F2"):
            self.generate_budget_button.disabled = True
        else:
            self.generate_budget_button.disabled = False
        
        self.page.update()
    
    def clean_data_table(self):
        for i in range(0,len(self.data_table.rows)):
            self.data_table.rows.remove(self.data_table.rows[0])
        
        #Clear order list
        self.order.clear()

        #Update price variables
        self.A_SUBTOTAL = 0
        self.N_SUBTOTAL = 0
        self.TOTAL_PRICE = self.A_SUBTOTAL + self.N_SUBTOTAL
        self.update_price_widgets()
        
        self.page.update()

    def get_product(self, mode, param):
        conn = sqlite3.connect(DATABASE_PATH)
        if(mode == "Product"):
            query = f"SELECT Codigo,Descripcion,Precio FROM Producto WHERE Descripcion = '{param}'"
        else:
             query = f"SELECT Codigo,Descripcion,Precio FROM Producto WHERE Codigo = '{param}'"

        try:
            cursor = conn.execute(query)
            record = cursor.fetchall()
            record = record[0]

            code = record[0]
            description = record[1]
            price = record[2]

            return code,description,price
        except:
            show_invalid_code_message(self.page)
            return "not exists","",""
        finally:
            conn.close()

    def update_mode_add_product(self, quantity,description):
        match(self.mode.value):
            case "F1":
                cant_A = quantity
                cant_N = 0
                cant_Total = cant_A
            case "F2":
                cant_A = math.floor(int(quantity) / 2)
                cant_N = math.ceil(int(quantity) / 2)
                cant_Total = quantity
            case "F3":
                cant_A = 0
                cant_N = quantity
                cant_Total = cant_N
            
        #Get product from DB
        code,product,price = self.get_product("Product",description)

        price = round(float(price),2)
        A_price = float(cant_A) * price * IVA
        N_price = float(cant_N) * price

        A_price = round(A_price, 2)
        N_price = round(N_price, 2)

        #Update total price variables
        self.A_SUBTOTAL += A_price
        self.N_SUBTOTAL += N_price
        self.TOTAL_PRICE += round(A_price + N_price,2)

        #Format prices
        price = '{:,.2f}'.format(price)
        A_price = '{:,.2f}'.format(A_price)
        N_price = '{:,.2f}'.format(N_price)


        #Insert into table
        self.insert_table_row(cant_A, cant_N, cant_Total, code, product, price, A_price, N_price)

        self.code
        self.update_price_widgets()
        self.page.update()

    def insert_table_row(self, cant_A, cant_N, cant_Total, code, product, price, A_price, N_price):
        self.data_table.rows.insert(
                len(self.data_table.rows),
                 ft.DataRow(
                            on_select_changed = lambda e: self.on_select_changed(e),
                            selected = False,
                            cells = [
                                ft.DataCell(ft.Text(cant_A,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(cant_N,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(cant_Total,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(str(code),text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(str(product),text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(price,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(A_price,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(N_price,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Row([ft.IconButton(tooltip = "Eliminar",
                                                                          icon = ft.icons.DELETE,
                                                                          icon_color = RED,
                                                                          on_click = lambda e: self.delete_row(e),
                                                                          data=len(self.data_table.rows)),
                                                    ft.IconButton(tooltip = "Editar",
                                                                          icon = ft.icons.EDIT,
                                                                          icon_color = "blue",
                                                                          on_click = lambda e: self.edit_row(e),
                                                                          data=len(self.data_table.rows),
                                                                          visible = True if self.current_checkbox_selection == "" else False)],
                                                                          alignment=ft.CrossAxisAlignment.CENTER))
                                ]
                 )
            )
        #If it is not a special product (not selected any checkbox)
        if(price != 0):
            #Add to orders list:
            self.order.append(product)

    def clear_fields(self):
        self.product.value = ""
        self.code.value = ""
        self.quantity.value = ""
        self.reset_checkboxs()
        self.page.update()

    def on_select_changed(self, e):
        if e.control.selected:
            e.control.selected = False
        else:
            e.control.selected = True

        self.page.update()
        
    def edit_row(self, e):
        selected_item = self.data_table.rows[e.control.data]

        #Enable quantity if disabled
        if(not self.quantity.error_text == ""):
            self.quantity.read_only = False
            self.quantity.error_text = ""

        self.data_table.rows.remove(selected_item)

        # GET ALL DATA TABLE ROWS
        all_rows = []
        special_products = []
        for row in self.data_table.rows:
            row_data = []
            for cell in row.cells:                
                try:
                    print(cell.content.value)
                    row_data.append(cell.content.value)
                except Exception as e:
                    pass
            #If not a special product, append to all rows
            if row_data[UNITARY_PRICE_IDX] != 0:
                all_rows.append(row_data)
            else:
                special_products.append(row_data)
        
        #DELETE ALL ROWS
        self.clean_data_table()        
        #UPDATE TABLE WITH NEW MODE
        for row in all_rows:
            self.update_mode_add_product(row[CANT_TOTAL_IDX],row[PRODUCT_IDX])

        if len(special_products) > 0:
            for row in special_products:
                self.current_checkbox_selection = "aux"
                self.insert_table_row(row[CANT_A_IDX],row[CANT_N_IDX],row[CANT_TOTAL_IDX],row[CODE_IDX],row[PRODUCT_IDX],
                                      row[UNITARY_PRICE_IDX],row[A_PRICE_IDX],row[N_PRICE_IDX])
                self.current_checkbox_selection = ""
        

        if(self.search_mode == "Product"):
            self.product.value = selected_item.cells[PRODUCT_IDX].content.value
        else:
            self.code.value = selected_item.cells[CODE_IDX].content.value
        
        self.quantity.value = selected_item.cells[CANT_TOTAL_IDX].content.value

        self.add_button.disabled = True
        self.update_button.disabled = False
        self.page.update()
    
    def delete_row(self, e):
        selected_item = self.data_table.rows[e.control.data]
        
        #Remove from order list
        product = selected_item.cells[PRODUCT_IDX].content.value
        #If it ends with "]" its a special product (any checkbox selected). So it isn't in order[].
        if(not product.endswith("]")):
            self.order.remove(product)

        #Remove from table
        self.data_table.rows.remove(selected_item)


        # GET ALL DATA TABLE ROWS
        all_rows = []
        special_products = []
        for row in self.data_table.rows:
            row_data = []
            for cell in row.cells:                
                try:
                    row_data.append(cell.content.value)
                except Exception as e:
                    pass
            #If not a special product, append to all rows
            if row_data[UNITARY_PRICE_IDX] != 0:
                all_rows.append(row_data)
            else:
                special_products.append(row_data)
        
        #DELETE ALL ROWS
        self.clean_data_table()        
        #UPDATE TABLE WITH NEW MODE
        for row in all_rows:
            self.update_mode_add_product(row[CANT_TOTAL_IDX],row[PRODUCT_IDX])
        
        if len(special_products) > 0:
            for row in special_products:
                self.current_checkbox_selection = "aux"
                self.insert_table_row(row[CANT_A_IDX],row[CANT_N_IDX],row[CANT_TOTAL_IDX],row[CODE_IDX],row[PRODUCT_IDX],
                                      row[UNITARY_PRICE_IDX],row[A_PRICE_IDX],row[N_PRICE_IDX])
                self.current_checkbox_selection = ""

        show_deleted_message(self.page)
        
        if(len(self.data_table.rows) == 0):
            self.generate_order_button.disabled = True
            self.generate_budget_button.disabled = True

        self.page.update()

    def set_header(self, e):
        self.hand_icon.visible = True
        self.welcome.value = f"Bienvenido ¡{self.user.value}!"
        self.page.update()

    def update_price_widgets(self):

        #Format prices to display
        total_price = '{:,.2f}'.format(self.TOTAL_PRICE)
        A_price = '{:,.2f}'.format(self.A_SUBTOTAL)
        N_price = '{:,.2f}'.format(self.N_SUBTOTAL)



        self.A_SUBTOTAL_text.value = f"Precio en A: ${A_price}"
        self.N_SUBTOTAL_text.value = f"Precio en N: ${N_price}"
        self.TOTAL_PRICE_text.value = f"Precio Total: ${total_price}"

        self.page.update()

    def generate_pdf(self, e):
        if not (self.pick_pdf_path_dialog.result.path is None):
            show_generating_pdf_message(self.page)

            date = datetime.now().strftime("%d-%m-%Y")
            filename = f"{e.path}\Presupuesto {self.client.value} {date}.pdf"
            pdf = FPDF(orientation='P', unit='mm', format='A4')

            pdf.set_margins(left=0,top=0,right=0)

            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 12)
            pdf.image(resource_path("assets/fenix_logo.png"),x = 10, y=7,w=30, h=22)

            #DATE
            pdf.cell(w=0, h=10, text=date, align='R',new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            #TITLE
            pdf.cell(w=0, h=20, text="DISTRIBUIDORA SANITARIA FENIX", align='C',new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            #CLIENT
            pdf.cell(w=0, h=10, text=self.client.value, align='C',new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            

            #TABLE HEADER
            pdf.cell(w=15, h=5,text="CANT.", align='L')
            pdf.cell(w=25, h=5,text="CÓDIGO", align='L')
            pdf.cell(w=100, h=5,text="ARTICULO", align='L')
            pdf.cell(w=30, h=5,text="PRECIO U.", align='L')
            pdf.cell(w=40, h=5,text="TOTAL", align='L',new_x=XPos.LMARGIN, new_y=YPos.NEXT)


            #CHOOSE MODE
            if(self.mode.value == "F1"):
                cant = CANT_A_IDX
                total = A_PRICE_IDX
            elif(self.mode.value == "F3"):
                cant = CANT_N_IDX
                total = N_PRICE_IDX

            #TABLE DATA

            #Remove actions from each row
            all_rows = []
            for row in self.data_table.rows:
                row_data = []
                for cell in row.cells:                
                    try:
                        row_data.append(cell.content.value)
                    except Exception as e:
                        pass

                all_rows.append(row_data)

            #Add table data
            pdf.set_font('Helvetica', '', 12)
            for row in all_rows:
                pdf.cell(w=15, h=5,text=str(row[cant]), align='L')
                pdf.cell(w=25, h=5,text=str(row[CODE_IDX]), align='L')
                pdf.cell(w=100, h=5,text=str(row[PRODUCT_IDX]), align='L')
                pdf.cell(w=30, h=5,text=f"${str(row[UNITARY_PRICE_IDX])}", align='L')
                pdf.cell(w=40, h=5,text=f"${str(row[total])}", align='L', new_x=XPos.LMARGIN, new_y=YPos.NEXT)



            #TOTAL PRICE
            pdf.set_font('Helvetica', 'B', 12)
            pdf.cell(w=15, h=10,text="", align='L')
            pdf.cell(w=25, h=10,text="", align='L')
            pdf.cell(w=100, h=10,text="", align='L')
            pdf.cell(w=30, h=10,text="TOTAL:", align='L')
            pdf.cell(w=40, h=10,text=f"${str('{:,.2f}'.format(self.TOTAL_PRICE))}", align='L', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.output(resource_path(filename))

            show_created_budget_message(self.page)

            self.clear_screen()
            self.page.update()

    def generate_excel(self,e):
        global PENDING_FILENAME
        filename = self.set_filename(e)
        if(not filename == ""):
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            header = self.generate_excel_header()
            for row in header:
                sheet.append(row)

                    
            #Bold header
            for i in range(1,3):
                for cell in sheet[i]:
                    cell.font = Font(bold=True)

            # Agregar encabezados de tabla
            sheet.append(table_headers)


            #       Add pending
            if not PENDING_FILENAME == "":
                pending_rows = self.get_pending_file_rows()
                for row in pending_rows:
                    sheet.append(row)
            
            #Bold pendings:
                for i in range(4, len(pending_rows)+4+1):
                    for cell in sheet[i]:
                        cell.font = Font(bold=True)
                

            #       Add order
            #Remove actions from each row
            all_rows = []
            for row in self.data_table.rows:
                row_data = []
                for cell in row.cells:                
                    try:
                        row_data.append(cell.content.value)
                    except Exception as e:
                        pass

                all_rows.append(row_data)

            # Agregar los datos a la hoja de cálculo
            for row in all_rows:
                sheet.append(row)

            min_row = len(header) + 1
            if PENDING_FILENAME == "":
                max_row = len(all_rows) + min_row
            else:
                max_row = len(all_rows) + min_row + len(pending_rows) + 1

            for i in range(min_row,max_row+1):
                column = 0
                for cell in sheet[i]:
                    if column == CANT_TOTAL_IDX:
                        cell.alignment = Alignment(horizontal="right")
                        cell.font = Font(bold=True)
                    else:
                        cell.alignment = Alignment(horizontal="left")
                    
                    column += 1
            #Dejar espacios:
            spaces = 2
            sheet.append([])
            sheet.append([])

            #Añadir importes totales:
            sheet.append(["Total en A:",self.A_SUBTOTAL])
            sheet.append(["Total en N:",self.N_SUBTOTAL])
            sheet.append(["Total del pedido:",self.TOTAL_PRICE])

            #Añadir formato de dinero
            sheet[f"B{max_row+spaces+1}"].number_format = '"$"#,##0.00_);("$"#,##0.00)'
            sheet[f"B{max_row+spaces+2}"].number_format = '"$"#,##0.00_);("$"#,##0.00)'
            sheet[f"B{max_row+spaces+3}"].number_format = '"$"#,##0.00_);("$"#,##0.00)'

            #Ajustar tamaño de columnas
            self.adjust_column_widths(sheet)
            # # Guardar el archivo Excel
            workbook.save(resource_path(filename))

            show_created_order_message(self.page)

            self.clear_screen()
            PENDING_FILENAME = ""
            self.page.update()

    def generate_excel_header(self):
        header = []
        user_row = []
        client_row = []

        #Leave 3 first cells in blank
        for i in range(0,3):
            user_row.append("")
            client_row.append("")

        user_row.append(self.mode.value)
        user_row.append(CURRENT_USER)
        
        #Get client time:
        conn = sqlite3.connect(DATABASE_PATH)
        query = f"SELECT Horario FROM Cliente WHERE Cliente = '{self.client.value}'"
        try:
            cursor = conn.execute(query)
            client_time = cursor.fetchall()
            client_time = client_time[0]
            
            
            client_row.append(client_time[0])
            client_row.append(f"{self.client.value} [{self.zone.value}]")
            
            header.append(user_row)
            header.append(client_row)

            return header
        
        except:
            show_error_message(self.page)
        finally:
            conn.close()

    def adjust_column_widths(self,sheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def set_filename(self, e):
        if(self.first_order_chkbox.value == True):
            if not (self.pick_excel_path_dialog.result.path is None):
                return self.get_excel_filename(e.path)
            else:
                return ""
        else:
            if not (self.pick_excel_filename_dialog.result.path is None):
                return f"{self.pick_excel_filename_dialog.result.path}.xlsx" 
            else:
                return ""

    def get_excel_filename(self, initial_path):
        date = datetime.now().strftime("%d-%m-%Y")
        return f"{initial_path}\{self.client.value} {date}.xlsx"

    def clear_screen(self):
        self.client.value = ""
        self.first_order_chkbox.label = ""
        self.first_order_chkbox.value = True
        self.zone.value = ""
        self.A_SUBTOTAL = 0
        self.N_SUBTOTAL = 0
        self.TOTAL_PRICE = 0
        self.generate_order_button.disabled = True
        self.generate_budget_button.disabled = True
        self.pending_file.value = ""
        self.pending_file.visible = False
        self.delete_pending_file.visible = False
        self.clean_data_table()
        self.clear_fields()
        self.update_price_widgets()

    def update_price_list(self,e: ft.FilePickerResultEvent):
        
        if e.files:
            selected_file = e.files[0]
            filename = selected_file.path
            
            show_updating_list_message(self.page)

            #Read xlsx file
            raw_product_list = pandas.read_excel(filename)

            #Replace NaN values (empty cells for 0)
            raw_product_list['Unnamed: 0'].fillna(0, inplace=True)

            #Get only the columns that we need
            codes = raw_product_list['Unnamed: 0'].values
            products_aux = raw_product_list['Unnamed: 1'].values
            prices = raw_product_list['Unnamed: 3'].values

            #Database connection
            conn = sqlite3.connect(resource_path(DATABASE_PATH))
            #Delete previous products
            query = f"DELETE FROM Producto"
            try:
                cursor = conn.execute(query)

                #Auxiliary index
                i = 0
                #Insert into database
                for product in products_aux:
                    #Avoid headers and blank cells
                    if not (codes[i] == 0 or codes[i] == "CODIGO"):
                        #Replace special characters
                        product = str(product).replace("'","''")
                        product = str(product).replace("*","")
                        product = str(product).replace("°","")
                        product = str(product).replace("ª","")
                        query = f"INSERT INTO Producto (Codigo,Descripcion,Precio) VALUES ('{codes[i]}','{product}',{prices[i]})"

                        try:
                            cursor = conn.execute(query)
                            conn.commit()
                            #products.append(product)
                        except sqlite3.Error as er:
                            show_error_message(self.page)
                        
                        
                    i += 1
                
                # self.snack_bar.open = False
                self.page.update()
                
                set_products(self.page)
                show_successfully_updated_list_message(self.page)

            except:
                show_error_message(self.page)
            finally:
                conn.close()

    def add_pending(self,e):
        global PENDING_FILENAME
        if e.files:
            PENDING_FILENAME = e.files[0].path
            show_added_pending_message(self.page) 
            self.pending_file.value = f"Archivo de pendientes: '{PENDING_FILENAME}'"
            self.pending_file.visible = True
            self.delete_pending_file.visible = True
            self.page.update()
    
    def delete_pending(self, e):
        global PENDING_FILENAME
        self.pending_file.visible = False
        self.pending_file.value = ""
        self.delete_pending_file.visible = False
        PENDING_FILENAME = ""
        self.page.update()

    def get_pending_file_rows(self):
        global PENDING_FILENAME
        pending_file = pandas.read_excel(PENDING_FILENAME)

        a_column = pending_file.columns[0]
        n_column = pending_file.columns[1]
        cant_total_column = pending_file.columns[2]
        code_column = pending_file.columns[3]
        desc_column = pending_file.columns[4]
        price_column = pending_file.columns[5]
        total_a_column = pending_file.columns[6]
        total_n_column = pending_file.columns[7]

        #Replace NaN values (empty cells for 0)
        pending_file[cant_total_column].fillna(0, inplace=True)

        #Get only the columns that we need
        A = pending_file[a_column].values
        N = pending_file[n_column].values
        cant_total = pending_file[cant_total_column].values
        code = pending_file[code_column].values
        desc = pending_file[desc_column].values
        price = pending_file[price_column].values
        total_a = pending_file[total_a_column].values
        total_n = pending_file[total_n_column].values

        rows = []
        i=2
        for i in range(2,len(cant_total)+1):
            if cant_total[i] == 0:
                break
            else:
                row = self.update_mode_pending_product(A[i],N[i],cant_total[i],code[i],desc[i],price[i],total_a[i],total_n[i])
                rows.append(row)
            i += 1

        return rows

    def update_mode_pending_product(self,A,N,CANT,CODE,DESC,PRICE,TOTAL_A,TOTAL_N):
        PRICE = str(PRICE).replace(",","")

        match(self.mode.value):
            case "F1":
                cant_total = int(A) + int(N)
                cant_a = cant_total
                cant_n = 0
                code = CODE
                desc = DESC
                price = PRICE
                total_a = float(PRICE) * 1.21
                total_n = 0
                
                return [cant_a,cant_n,cant_total,code,desc,price,total_a,total_n]


            case "F2":
                cant_total = int(A) + int(N)
                cant_a = math.floor(int(cant_total) / 2)
                cant_n = math.ceil(int(cant_total) / 2)
                code = CODE
                desc = DESC
                price = PRICE
                total_a = float(PRICE) *cant_a * 1.21
                total_n = float(PRICE) * cant_n

                return [cant_a,cant_n,cant_total,code,desc,price,total_a,total_n]
                
            case "F3":
                cant_total = int(A) + int(N)
                cant_a = 0
                cant_n = cant_total
                code = CODE
                desc = DESC
                price = PRICE
                total_n = PRICE
                total_a = 0

                return [cant_a,cant_n,cant_total,code,desc,price,total_a,total_n]

    def set_user(self,e):
        global CURRENT_USER
        self.client.value = ""
        self.zone.value = ""
        self.client.read_only = False
        self.generate_order_button.disabled = True
        self.generate_budget_button.disabled = True
        self.client_search_list.visible = False
        self.mode.visible = True
        self.client.error_text = ""
        set_clients(self.user.value, self.page)
        self.welcome_title.value = f"¡Bienvenido {self.user.value}!"
        self.page.update()

        CURRENT_USER = self.user.value

    def go_to_clients_screen(self, e):
        global CURRENT_USER
        self.page.views.clear()
        self.page.views.append(Clients(self.page).build())
        CURRENT_USER = ""
        self.page.update()

class Clients(ft.Container):
    def __init__(self, page: ft.Page):
        super().__init__(expand = True)

        self.is_updating = False
        self.page = page
        self.create_widgets()

     
        #FORM
        self.client_form = ft.Container(bgcolor = BG_COLOR,
                                 border_radius = 10,
                                 padding = 10,
                                 col = 3,
                                content = ft.Column(
                                    controls = [
                                        self.title,
                                        self.client,
                                        self.zone,
                                        self.times,
                                        ft.Row(
                                            controls = [
                                                self.add_button,
                                                self.update_button
                                                ],
                                                alignment=ft.MainAxisAlignment.CENTER
                                            )
                                        
                                    ])
                                )

        #TABLE
        self.table = ft.Container(bgcolor = BG_COLOR,
                                  border_radius = 10,
                                  col = 9,
                                  content = 
                                            ft.Column(
                                                expand = True,
                                                scroll = "auto",
                                                controls = [
                                                    ft.ResponsiveRow([
                                                        self.data_table
                                                    ]
                                                    )
                                                ]
                                            )
                                    )         


        self.screen = ft.Container(bgcolor = BG_COLOR,
                                   expand=True,
                                    content=ft.Column(
                                        controls = [
                                            ft.Container(content = ft.ResponsiveRow(controls=[self.client_form,self.table]),expand=True)
                                        ]))

        self.content = ft.View(
                "/clients_screen",
                [
                    self.appbar,
                    self.screen,
                ],
                bgcolor=BG_COLOR                
            )

    def build(self):
        return self.content
    
    def create_widgets(self):

        # APPBAR WIDGETS
        self.user = ft.Dropdown(label = "Usuario",
                        text_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                        label_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                       suffix_icon=ft.icons.PERSON,
                       options = set_dropdown_options(users_list),
                       border_color = BUTTONS_BORDER_COLOR,
                       bgcolor=DROPDOWN_BG_COLOR,
                       fill_color=BG_COLOR,
                       on_change = self.set_user)
        
        self.welcome_title = ft.Text("SELECCIONE USUARIO", font_family=FONT,size=30,color="white")

        self.fenix_logo = ft.Image(src=resource_path("assets\\fenix_logo.png"),height=60,width=60,color=RED)

        self.appbar = ft.AppBar(
        leading=ft.Icon(ft.icons.PEOPLE, color=RED),
        leading_width=40,
        center_title=True,
        title=ft.Container(ft.Row(controls=[self.fenix_logo,self.welcome_title],
                                  alignment=ft.MainAxisAlignment.CENTER)),
                                  bgcolor=BG_COLOR,
                                  actions=[
                                      self.user,
                                        ft.IconButton(ft.icons.HOME,tooltip="Inicio",
                                                    on_click=self.go_to_home_screen,
                                                    icon_color=HOME_ICON_COLOR)
                                            ]
                                        )

        #FORM WIDGETS
        self.title = ft.Container(ft.Text("Datos de cliente",text_align=ft.TextAlign.CENTER, size = 30,
                                          font_family=FONT,color="white"),
                                          alignment=ft.alignment.center)

        self.client = ft.TextField(label = "Cliente",
                                   text_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                                   label_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                                   cursor_color=CURSOR_COLOR,
                                   border_color = BUTTONS_BORDER_COLOR,
                                   read_only=True,
                                   on_focus=self.check_user,
                                   on_change=self.check_client_change,
                                   on_submit=lambda _:self.zone.focus())
        
        self.zone = ft.TextField(label = "Zona",
                                 text_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                                 label_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                                 cursor_color=CURSOR_COLOR,
                                 border_color = BUTTONS_BORDER_COLOR,
                                 read_only=True,
                                 on_focus=self.check_client,
                                 on_change=self.check_zone_change,
                                 on_submit=lambda _:self.times.focus())
        
        self.times = ft.TextField(label = "Horario",
                                  text_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                                  label_style=ft.TextStyle(font_family=FONT,size=20,color="white"),
                                  border_color = BUTTONS_BORDER_COLOR,
                                  cursor_color=RED,
                                  helper_text="Ej: 8a12hs - 14a19hs (En caso de no saber -)",
                                  helper_style=ft.TextStyle(font_family=FONT,color="white"),
                                  on_change=self.check_times_input,
                                  on_focus=self.check_zone,
                                  on_submit=self.check_submit)

        self.add_button = ft.ElevatedButton(text = "Añadir", icon=ft.icons.ADD,on_click = self.add_client,disabled=True,
                                            style=ft.ButtonStyle(color={"":ADD_BUTTON_COLOR,"disabled":BG_COLOR},
                                                                 bgcolor={"":ADD_BUTTON_BGCOLOR,"disabled":DISABLED_GREY}))

        self.update_button = ft.ElevatedButton(text = "Actualizar", icon=ft.icons.REFRESH,disabled=True,on_click = self.update_client,
                                               style=ft.ButtonStyle(color={"":ADD_BUTTON_COLOR,"disabled":BG_COLOR},
                                                                 bgcolor={"":ADD_BUTTON_BGCOLOR,"disabled":DISABLED_GREY}))

        #TABLE WIDGETS
        self.data_table = ft.DataTable(expand = True,
                                       border = ft.border.all(2,BUTTONS_BORDER_COLOR),
                                       data_row_color = {ft.MaterialState.SELECTED: BUTTONS_BORDER_COLOR,
                                                         ft.MaterialState.PRESSED: "black"},
                                        column_spacing=10,
                                        border_radius = 10,
                                        columns = [
                                             ft.DataColumn(label=ft.Container(ft.Text("Cliente",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=150)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Zona",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=150)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Horario",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT), width=150)),
                                             ft.DataColumn(label=ft.Container(ft.Text("Acciones",color = BUTTONS_BORDER_COLOR, weight = "bold",text_align=ft.TextAlign.CENTER,font_family=FONT),
                                                                              width=100,expand=True,alignment=ft.alignment.center))
                                        ],
                                        rows = [])

        #DIALOG
        self.confirmation_dialog = ft.AlertDialog(
            modal=True,
            bgcolor=RED,
            content_text_style=ft.TextStyle(font_family=FONT,color="white"),
            title_text_style=ft.TextStyle(font_family=FONT,color="white"),
            title=ft.Text("ELIMINANDO CLIENTE",font_family=FONT),
            content=ft.Text("¿Está seguro que desea eliminar el cliente?",font_family=FONT),
            actions=[
                ft.ElevatedButton(text = "Sí",on_click = self.handle_dialog,
                                  color=ADD_BUTTON_COLOR,bgcolor=ADD_BUTTON_BGCOLOR),
                ft.ElevatedButton(text = "No",on_click = self.handle_dialog,
                                  color=ADD_BUTTON_COLOR,bgcolor=ADD_BUTTON_BGCOLOR)
            ],
            actions_alignment=ft.MainAxisAlignment.CENTER)

    def check_times_input(self, e):
        if(e.control.value != "" and self.client.value != "" and self.zone.value != ""):
            if(self.is_updating):
                self.update_button.disabled = False
            else:
                self.add_button.disabled = False
        else:
            if(self.is_updating):
                self.update_button.disabled = True
            else:
                self.add_button.disabled = True

        self.page.update()

    def check_client_change(self, e):
        self.client.value = e.control.value.upper()
        self.client.update()

        if(e.control.value == ""):
            self.zone.read_only = True
        else:
            self.zone.read_only = False

        self.zone.error_text = ""
        
        self.zone.update()

    def check_user(self, e):
        if(CURRENT_USER == ""):
            self.client.error_text = "Primero debe seleccionar usuario" 
        else:
            self.client.error_text = "" 

        if(self.client.value == ""):
            self.zone.read_only = True
        else:
            self.zone.read_only = True

        self.client.update()
        self.zone.update()

    def check_client(self, e):
        if(self.client.value == ""):
            self.zone.read_only = True
            self.zone.error_text = "Primero debe ingresar el cliente"
        else:
            self.zone.read_only = False
            self.zone.error_text = ""        

        self.zone.update()
        self.client.update()

    def check_zone_change(self, e):
        self.zone.value = e.control.value.upper()
        self.zone.update()
        if(e.control.value == ""):
            self.times.read_only = True
        else:
            self.times.read_only = False

        self.times.error_text = ""
        
        self.times.update()

    def check_zone(self, e):
        if(self.zone.value == ""):
            self.times.read_only = True
            self.times.error_text = "Primero debe ingresar la zona"
        else:
            self.times.read_only = False
            self.times.error_text = ""        

        self.times.update()

    def check_submit(self, e):
        if(self.add_button.disabled == True):
            self.update_client(e)
        elif(self.update_button.disabled == True):
            self.add_client(e)

    def set_user(self, e):
        global CURRENT_USER
        CURRENT_USER = self.user.value
        self.welcome_title.value = "Gestión de clientes"
        self.client.read_only = False
        self.client.error_text = ""
        self.user.update()

        self.clean_data_table()
        self.fill_table()

        self.page.update()

    def add_client(self, e):
        if(self.client.value == "" or self.zone.value == "" or self.times.value == ""):
            show_empty_fields_message(self.page)
        else:
            conn = sqlite3.connect(DATABASE_PATH)

            #GET USER ID
            query = f"SELECT ID FROM Usuario WHERE Usuario = '{self.user.value}'"
            try:
                cursor = conn.execute(query)
                id = cursor.fetchone()

                #INSERT CLIENT
                query = f"INSERT INTO Cliente (Cliente,Zona,Horario,Usuario) VALUES ('{self.client.value}','{self.zone.value}','{self.times.value}',{id[0]})"
                try:
                    cursor = conn.execute(query)
                    conn.commit()
                    self.insert_table_row(self.client.value, self.zone.value, self.times.value)

                    show_added_client_message(self.page)

                    self.clear_fields()
                except:
                    show_error_message(self.page)
                finally:
                    self.page.update()
            except:
                show_error_message(self.page)
            finally:
                conn.close()

    def update_client(self, e):
        if(self.client.value == "" or self.zone.value == "" or self.times.value == ""):
            show_empty_fields_message(self.page)
        else:
            conn = sqlite3.connect(DATABASE_PATH)

            #UPDATE CLIENT
            query = f"UPDATE Cliente SET Cliente = '{self.client.value}', Zona = '{self.zone.value}', Horario = '{self.times.value}' WHERE ID = {self.selected_client}"
            try:
                cursor = conn.execute(query)
                self.insert_table_row(self.client.value, self.zone.value, self.times.value)

                show_updated_client_message(self.page)

                conn.commit()
                self.clear_fields()
            except:
                show_error_message(self.page)

            finally:
                self.data_table.disabled = False
                conn.close()


        self.add_button.disabled = False
        self.update_button.disabled = True
        self.is_updating = False
        self.page.update()

    def fill_table(self):
        conn = sqlite3.connect(DATABASE_PATH)
        query = f"SELECT Cliente,Zona,Horario FROM Cliente INNER JOIN Usuario ON Cliente.Usuario = Usuario.ID WHERE Usuario.Usuario = '{self.user.value}' ORDER BY Zona DESC"
        try:
            cursor = conn.execute(query)
            clients = cursor.fetchall()

            for client in clients:
                self.insert_table_row(client[CLIENT_IDX],client[ZONE_IDX],client[TIME_IDX])
            

            self.page.update()
        except:
            self.page.update()
        finally:
            conn.close()

    def insert_table_row(self, client, zone, time):
        self.data_table.rows.insert(
                len(self.data_table.rows),
                 ft.DataRow(
                            on_select_changed = lambda e: e,
                            selected = False,
                            cells = [
                                ft.DataCell(ft.Text(client,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(zone,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Text(time,text_align=ft.TextAlign.CENTER,font_family=FONT,color="white")),
                                ft.DataCell(ft.Row([ft.IconButton(tooltip = "Eliminar",
                                                                          icon = ft.icons.DELETE,
                                                                          icon_color = RED,
                                                                          on_click = lambda e: self.delete_row(e),
                                                                          data=len(self.data_table.rows)),
                                                    ft.IconButton(tooltip = "Editar",
                                                                          icon = ft.icons.EDIT,
                                                                          icon_color = "blue",
                                                                          on_click = lambda e: self.edit_row(e),
                                                                          data=len(self.data_table.rows))],
                                                                          alignment=ft.CrossAxisAlignment.CENTER))
                                ]
                 )
            )

    def edit_row(self, e):
        #Disable table to prevent picking another record
        self.data_table.disabled = True
        self.data_table.update()

        selected_item = self.data_table.rows[e.control.data]

        #Set selected item's ID
        conn = sqlite3.connect(DATABASE_PATH)
        query = f"SELECT ID FROM Cliente WHERE Cliente = '{selected_item.cells[CLIENT_IDX].content.value}' AND Zona = '{selected_item.cells[ZONE_IDX].content.value}'"
        try:
            cursor = conn.execute(query)
            client_id = cursor.fetchone()

            self.selected_client = client_id[0]

            self.data_table.rows.remove(selected_item)

            # GET ALL DATA TABLE ROWS
            all_rows = []
            for row in self.data_table.rows:
                row_data = []
                for cell in row.cells:                
                    try:
                        row_data.append(cell.content.value)
                    except Exception as e:
                        pass

                all_rows.append(row_data)
            
            #DELETE ALL ROWS
            self.clean_data_table()        
            #UPDATE TABLE WITH NEW MODE
            for row in all_rows:
                self.insert_table_row(row[CLIENT_IDX],row[ZONE_IDX],row[TIME_IDX])

            self.client.value = selected_item.cells[CLIENT_IDX].content.value
            self.zone.value = selected_item.cells[ZONE_IDX].content.value
            self.times.value = selected_item.cells[TIME_IDX].content.value

            self.add_button.disabled = True
            self.update_button.disabled = False
            self.is_updating = True
            self.page.update()
        except:
            show_error_message(self.page)
        finally:
            conn.close()
    
    def delete_row(self, e):
        self.selected_item = self.data_table.rows[e.control.data]

        #Set selected item's ID
        conn = sqlite3.connect(DATABASE_PATH)
        query = f"SELECT ID FROM Cliente WHERE Cliente = '{self.selected_item.cells[CLIENT_IDX].content.value}' AND Zona = '{self.selected_item.cells[ZONE_IDX].content.value}'"
        try:
            cursor = conn.execute(query)
            client_id = cursor.fetchone()

            self.selected_client = client_id[0]

            self.page.open(self.confirmation_dialog)
        except:
            show_error_message(self.page)
        finally:
            conn.close()

    def clean_data_table(self):
        for i in range(0,len(self.data_table.rows)):
            self.data_table.rows.remove(self.data_table.rows[0])
        
        self.page.update()

    def clear_fields(self):
        self.client.value = ""
        self.zone.value = ""
        self.times.value = ""

        self.page.update()

    def handle_dialog(self, e):
        if(e.control.text == "No"):
            self.page.close(self.confirmation_dialog)
        elif(e.control.text == "Sí"):
            conn = sqlite3.connect(DATABASE_PATH)
            query = f"DELETE FROM Cliente WHERE ID = {self.selected_client}"

            try:
                cursor = conn.execute(query)
                conn.commit()
                #Remove from table
                self.data_table.rows.remove(self.selected_item)

                # GET ALL DATA TABLE ROWS
                all_rows = []
                for row in self.data_table.rows:
                    row_data = []
                    for cell in row.cells:                
                        try:
                            row_data.append(cell.content.value)
                        except Exception as e:
                            pass

                    all_rows.append(row_data)
                
                #DELETE ALL ROWS
                self.clean_data_table()     

                for row in all_rows:
                    self.insert_table_row(row[CLIENT_IDX],row[ZONE_IDX],row[TIME_IDX])

                self.page.close(self.confirmation_dialog)

                show_deleted_client_message(self.page)
            except e:
                self.page.close(self.confirmation_dialog)
                show_error_message(self.page)
            finally:
                conn.close()

        self.page.update()

    def go_to_home_screen(self, e):
        global CURRENT_USER
        CURRENT_USER = ""
        self.page.views.clear()
        form = Form(self.page)
        self.page.views.append(form.build())
        self.page.go("/home_screen")
        # self.page.views.append(Form(self.page).build())
        
        self.page.update()  

#MAIN

def main(page: ft.Page):
    page.theme_mode = "dark"
    page.window.icon = resource_path("assets/fenix_icon.ico")
    page.fonts = {
        "ModeRustic": resource_path("assets\\Fonts\\modeRustic\\static\\Moderustic-Regular.ttf")
    }
    set_users_list(page)
    form = Form(page)
    page.views.append(form.build())
    page.update()

ft.app(main)


